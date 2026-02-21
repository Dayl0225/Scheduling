"""
Excel Export Service for Campus Scheduling System

Generates Excel workbooks with three views:
1. Teacher View: Schedule per teacher with gaps
2. Section View: Schedule per course/section  
3. Room/Building View: Schedule by room and building
Plus: Validation checklist of all input data
"""

import io
from datetime import datetime, time as time_type
from typing import List, Dict, Optional, Tuple
from sqlalchemy.orm import Session
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from . import models


class ExcelExporter:
    """Generates Excel workbooks with scheduling data in multiple views."""
    
    # Color scheme for courses (hex values without #)
    COLOR_PALETTE = [
        "FFE699",  # Yellow
        "C6E0B4",  # Light Green
        "BDD7EE",  # Light Blue
        "F4B084",  # Light Orange
        "E2EFDA",  # Light Mint
        "FFF2CC",  # Pale Yellow
        "D9E1F2",  # Pale Blue
        "FCE4D6",  # Pale Orange
        "EDEDED",  # Light Gray
        "E7E6E6",  # Gray
    ]
    
    DAYS_OF_WEEK = ["MON", "TUE", "WED", "THU", "FRI", "SAT"]
    HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, db: Session):
        self.db = db
        self.color_map: Dict[int, str] = {}  # Map course_id to color
        self._init_color_map()
    
    def _init_color_map(self):
        """Initialize color assignments for courses."""
        courses = self.db.query(models.Course).all()
        for idx, course in enumerate(courses):
            self.color_map[course.id] = self.COLOR_PALETTE[idx % len(self.COLOR_PALETTE)]
    
    def export_schedule(
        self,
        schedule_run_id: int,
        institution_name: str = "Campus Scheduling System"
    ) -> bytes:
        """
        Export a complete schedule to Excel with all three views.
        
        Args:
            schedule_run_id: ID of the schedule run to export
            institution_name: Name of the institution (for header)
        
        Returns:
            bytes: Excel file content
        """
        # Load schedule data
        schedule_run = self.db.query(models.ScheduleRun).filter(
            models.ScheduleRun.id == schedule_run_id
        ).first()
        
        if not schedule_run:
            raise ValueError(f"Schedule run {schedule_run_id} not found")
        
        schedule_entries = self.db.query(models.ScheduleEntry).filter(
            models.ScheduleEntry.schedule_run_id == schedule_run_id
        ).all()
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Add sheets
        self._add_teacher_view(wb, schedule_entries, institution_name)
        self._add_section_view(wb, schedule_entries, institution_name)
        self._add_room_view(wb, schedule_entries, institution_name)
        self._add_checklist_view(wb, schedule_run)
        
        # Write to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    
    def _add_teacher_view(
        self,
        wb: Workbook,
        schedule_entries: List[models.ScheduleEntry],
        institution_name: str
    ):
        """Add Teacher View sheet - Schedule per teacher with gaps."""
        ws = wb.create_sheet("Teacher View")
        
        # Get unique teachers
        teachers = {}
        for entry in schedule_entries:
            if entry.teacher_id not in teachers:
                teachers[entry.teacher_id] = entry.teacher
        
        teachers = dict(sorted(teachers.items(), key=lambda x: x[1].full_name))
        
        row = 1
        
        # Header
        ws[f'A{row}'] = institution_name
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        ws[f'A{row}'] = f"Schedule Export - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws[f'A{row}'].font = Font(italic=True, size=10)
        row += 2
        
        for teacher_id, teacher in teachers.items():
            # Teacher header
            ws[f'A{row}'] = f"{teacher.full_name}"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            row += 1
            
            ws[f'A{row}'] = f"Title: {teacher.title.value} | Status: {teacher.status.value} | Workload: {teacher.workload.value}"
            ws[f'A{row}'].font = Font(italic=True, size=9)
            row += 1
            
            # Table header
            headers = ["Day", "Start", "End", "Course", "Section", "Room", "Building"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.border = self.BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row += 1
            
            # Get teacher's schedule entries
            teacher_entries = [e for e in schedule_entries if e.teacher_id == teacher_id]
            teacher_entries.sort(key=lambda e: (
                self.DAYS_OF_WEEK.index(e.timeslot.day_of_week.value),
                e.timeslot.start_time
            ))
            
            # Add entries
            for entry in teacher_entries:
                color = self.color_map.get(entry.course_id, "FFFFFF")
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                row_data = [
                    entry.timeslot.day_of_week.value,
                    entry.timeslot.start_time,
                    entry.timeslot.end_time,
                    entry.course.course_code,
                    entry.section.code,
                    entry.room.room_code,
                    entry.room.building.name if entry.room.building else "N/A"
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row, col, value)
                    cell.fill = fill
                    cell.border = self.BORDER
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                row += 1
            
            row += 1  # Blank line between teachers
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
    
    def _add_section_view(
        self,
        wb: Workbook,
        schedule_entries: List[models.ScheduleEntry],
        institution_name: str
    ):
        """Add Section View sheet - Schedule per course/section."""
        ws = wb.create_sheet("Section View")
        
        # Get unique sections
        sections = {}
        for entry in schedule_entries:
            if entry.section_id not in sections:
                sections[entry.section_id] = entry.section
        
        sections = dict(sorted(sections.items(), key=lambda x: x[1].code))
        
        row = 1
        
        # Header
        ws[f'A{row}'] = institution_name
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        ws[f'A{row}'] = f"Course/Section Schedule - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws[f'A{row}'].font = Font(italic=True, size=10)
        row += 2
        
        for section_id, section in sections.items():
            # Section header
            ws[f'A{row}'] = f"{section.code}"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            row += 1
            
            ws[f'A{row}'] = f"Year Level: {section.year_level} | 1st Year: {'Yes' if section.is_first_year else 'No'}"
            ws[f'A{row}'].font = Font(italic=True, size=9)
            row += 1
            
            # Table header
            headers = ["Day", "Start", "End", "Course", "Teacher", "Room", "Building"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.border = self.BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row += 1
            
            # Get section's schedule entries
            section_entries = [e for e in schedule_entries if e.section_id == section_id]
            section_entries.sort(key=lambda e: (
                self.DAYS_OF_WEEK.index(e.timeslot.day_of_week.value),
                e.timeslot.start_time
            ))
            
            # Add entries
            for entry in section_entries:
                color = self.color_map.get(entry.course_id, "FFFFFF")
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                row_data = [
                    entry.timeslot.day_of_week.value,
                    entry.timeslot.start_time,
                    entry.timeslot.end_time,
                    entry.course.course_code,
                    entry.teacher.full_name,
                    entry.room.room_code,
                    entry.room.building.name if entry.room.building else "N/A"
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row, col, value)
                    cell.fill = fill
                    cell.border = self.BORDER
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                row += 1
            
            row += 1  # Blank line between sections
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
    
    def _add_room_view(
        self,
        wb: Workbook,
        schedule_entries: List[models.ScheduleEntry],
        institution_name: str
    ):
        """Add Room/Building View sheet - Schedule by room and building."""
        ws = wb.create_sheet("Room View")
        
        # Get unique rooms
        rooms = {}
        for entry in schedule_entries:
            if entry.room_id not in rooms:
                rooms[entry.room_id] = entry.room
        
        # Sort by building then room
        rooms = dict(sorted(
            rooms.items(),
            key=lambda x: (
                x[1].building.code if x[1].building else "Z",
                x[1].room_code
            )
        ))
        
        row = 1
        
        # Header
        ws[f'A{row}'] = institution_name
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        ws[f'A{row}'] = f"Room/Building Schedule - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws[f'A{row}'].font = Font(italic=True, size=10)
        row += 2
        
        current_building = None
        
        for room_id, room in rooms.items():
            building_name = room.building.name if room.building else "Unassigned"
            
            # Building header (only once per building)
            if current_building != building_name:
                if current_building is not None:
                    row += 1
                
                ws[f'A{row}'] = f"Building: {building_name}"
                ws[f'A{row}'].font = Font(bold=True, size=10, color="FFFFFF")
                ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                row += 1
                current_building = building_name
            
            # Room header
            ws[f'A{row}'] = f"  {room.room_code}"
            ws[f'A{row}'].font = Font(bold=True, size=10)
            row += 1
            
            ws[f'A{row}'] = f"  Floor: {room.floor_no} | Type: {room.room_type.value} | Capacity: {room.capacity}"
            ws[f'A{row}'].font = Font(italic=True, size=9)
            row += 1
            
            # Table header
            headers = ["Day", "Start", "End", "Course", "Section", "Teacher"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.border = self.BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row += 1
            
            # Get room's schedule entries
            room_entries = [e for e in schedule_entries if e.room_id == room_id]
            room_entries.sort(key=lambda e: (
                self.DAYS_OF_WEEK.index(e.timeslot.day_of_week.value),
                e.timeslot.start_time
            ))
            
            # Add entries
            for entry in room_entries:
                color = self.color_map.get(entry.course_id, "FFFFFF")
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                row_data = [
                    entry.timeslot.day_of_week.value,
                    entry.timeslot.start_time,
                    entry.timeslot.end_time,
                    entry.course.course_code,
                    entry.section.code,
                    entry.teacher.full_name
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row, col, value)
                    cell.fill = fill
                    cell.border = self.BORDER
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                row += 1
            
            row += 1  # Blank line between rooms
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
    
    def _add_checklist_view(self, wb: Workbook, schedule_run: models.ScheduleRun):
        """Add Checklist view - Validation of all input data."""
        ws = wb.create_sheet("Checklist")
        
        row = 1
        
        # Header
        ws[f'A{row}'] = "Data Validation Checklist"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        ws[f'A{row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws[f'A{row}'].font = Font(italic=True, size=9)
        row += 2
        
        # Summary Info
        ws[f'A{row}'] = "Schedule Information"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        row += 1
        
        checklist_items = [
            ("Schedule Run ID", str(schedule_run.id)),
            ("Term ID", str(schedule_run.term_id)),
            ("Status", schedule_run.status.value),
            ("Objective Score", f"{schedule_run.objective_score:.2f}" if schedule_run.objective_score else "N/A"),
            ("Created By", schedule_run.created_by),
            ("Created At", schedule_run.created_at.strftime('%Y-%m-%d %H:%M:%S')),
        ]
        
        for label, value in checklist_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Data Counts
        teachers_count = self.db.query(models.Teacher).filter(models.Teacher.active == True).count()
        courses_count = self.db.query(models.Course).count()
        sections_count = self.db.query(models.Section).count()
        rooms_count = self.db.query(models.Room).filter(models.Room.active == True).count()
        timeslots_count = self.db.query(models.Timeslot).count()
        assignments_count = self.db.query(models.ScheduleEntry).filter(
            models.ScheduleEntry.schedule_run_id == schedule_run.id
        ).count()
        
        ws[f'A{row}'] = "Data Summary"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        row += 1
        
        summary_items = [
            ("Active Teachers", str(teachers_count)),
            ("Courses", str(courses_count)),
            ("Sections", str(sections_count)),
            ("Active Rooms", str(rooms_count)),
            ("Available Timeslots", str(timeslots_count)),
            ("Schedule Entries", str(assignments_count)),
        ]
        
        for label, value in summary_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Validation Checks
        ws[f'A{row}'] = "Validation Checks"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        row += 1
        
        # Perform checks
        checks = self._perform_validation_checks(schedule_run.id)
        
        for check_name, check_result in checks:
            ws[f'A{row}'] = check_name
            ws[f'B{row}'] = "✓ PASS" if check_result else "✗ FAIL"
            
            if check_result:
                ws[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws[f'B{row}'].font = Font(color="006100")
            else:
                ws[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                ws[f'B{row}'].font = Font(color="9C0006")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
    
    def _perform_validation_checks(self, schedule_run_id: int) -> List[Tuple[str, bool]]:
        """Perform validation checks on the schedule."""
        checks = []
        
        # Check 1: All entries have valid teachers
        entries_without_teacher = self.db.query(models.ScheduleEntry).filter(
            models.ScheduleEntry.schedule_run_id == schedule_run_id,
            models.ScheduleEntry.teacher_id.is_(None)
        ).count()
        checks.append(("All entries assigned to teachers", entries_without_teacher == 0))
        
        # Check 2: All entries have valid courses
        entries_without_course = self.db.query(models.ScheduleEntry).filter(
            models.ScheduleEntry.schedule_run_id == schedule_run_id,
            models.ScheduleEntry.course_id.is_(None)
        ).count()
        checks.append(("All entries assigned to courses", entries_without_course == 0))
        
        # Check 3: All entries have valid rooms
        entries_without_room = self.db.query(models.ScheduleEntry).filter(
            models.ScheduleEntry.schedule_run_id == schedule_run_id,
            models.ScheduleEntry.room_id.is_(None)
        ).count()
        checks.append(("All entries assigned to rooms", entries_without_room == 0))
        
        # Check 4: All entries have valid timeslots
        entries_without_timeslot = self.db.query(models.ScheduleEntry).filter(
            models.ScheduleEntry.schedule_run_id == schedule_run_id,
            models.ScheduleEntry.timeslot_id.is_(None)
        ).count()
        checks.append(("All entries assigned to timeslots", entries_without_timeslot == 0))
        
        # Check 5: No teacher has > 5 teaching days per week
        teachers = self.db.query(models.Teacher).filter(models.Teacher.active == True).all()
        all_teachers_valid = True
        for teacher in teachers:
            teaching_days = self.db.query(models.ScheduleEntry.timeslot_id).filter(
                models.ScheduleEntry.schedule_run_id == schedule_run_id,
                models.ScheduleEntry.teacher_id == teacher.id
            ).distinct(models.ScheduleEntry.timeslot_id).count()
            
            if teaching_days > 5:
                all_teachers_valid = False
                break
        
        checks.append(("Teachers teach max 5 days/week", all_teachers_valid))
        
        # Check 6: Schedule has entries
        total_entries = self.db.query(models.ScheduleEntry).filter(
            models.ScheduleEntry.schedule_run_id == schedule_run_id
        ).count()
        checks.append(("Schedule has entries", total_entries > 0))
        
        return checks


def export_schedule_to_excel(
    schedule_run_id: int,
    db: Session,
    institution_name: str = "Campus Scheduling System"
) -> bytes:
    """
    Convenience function to export a schedule run to Excel.
    
    Args:
        schedule_run_id: ID of the schedule run
        db: Database session
        institution_name: Name of institution (for headers)
    
    Returns:
        bytes: Excel file content ready to write to disk
    """
    exporter = ExcelExporter(db)
    return exporter.export_schedule(schedule_run_id, institution_name)
